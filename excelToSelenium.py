# EP.4 Excel Automation
from openpyxl import load_workbook
import os
import time
from selenium import webdriver

excelfile = load_workbook('../asset/wikipediatoexcel.xlsx')
sheet = excelfile.active

# print(sheet['B1'].value)

pdname = sheet.cell(row=3, column=2).value
pdprice = sheet.cell(row=3, column=3).value
pddetail = sheet.cell(row=3, column=4).value

# print(pdname, pdprice, pddetail)

alltitle, allprice, alldata = [], [], []

for i in range(3, 5):
    pdname = sheet.cell(row=i, column=2).value
    pdprice = sheet.cell(row=i, column=3).value
    pddetail = sheet.cell(row=i, column=4).value
    alltitle.append(pdname)
    allprice.append(pdprice)
    alldata.append(pddetail)

driverpath = '/Users/la/www/pythonBootCamp/code/chromedriver'
driver = webdriver.Chrome(driverpath)

url = 'http://uncle-machine.com/login/'
driver.get(url)

username = driver.find_element_by_id('username')
username.send_keys('pythonbootcamprpa@gmail.com')
password = driver.find_element_by_id('password')
password.send_keys('1234')
button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
button.click()
# password.send_keys(Keys.RETURN)

addurl = 'http://uncle-machine.com/addproduct/'
driver.get(addurl)

for pn, pp, pd in zip(alltitle, allprice, alldata):
    pdname = driver.find_element_by_id('name')
    pdprice = driver.find_element_by_id('price')
    pddetail = driver.find_element_by_id('detail')
    # photo = driver.find_element_by_id('photo')

    # pd1_name = pn
    # pd1_price = pp
    # pd1_detail = pd
    # photopath = pt

    pdname.send_keys(pn)
    pdprice.send_keys(pp)
    pddetail.send_keys(pd)
    # photo.send_keys(pt)

    time.sleep(3)

    button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
    button.click()
